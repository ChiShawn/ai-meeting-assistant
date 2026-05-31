#!/usr/bin/env python3
import io
import json
import os
import re
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

import torch
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from transformers import AutoModelForCausalLM, AutoTokenizer


LLM_MODEL_ID = os.getenv("LLM_MODEL_ID", "Qwen/Qwen3-0.6B")
WHISPER_MODEL_SIZE = os.getenv("WHISPER_MODEL_SIZE", "small")
WHISPER_DEVICE = os.getenv("WHISPER_DEVICE", "cuda")
WHISPER_COMPUTE_TYPE = os.getenv("WHISPER_COMPUTE_TYPE", "float16")
CORRECTION_DIR = Path(os.getenv("CORRECTION_DIR", str(Path.home() / "meeting_gpu_backend" / "corrections")))
CORRECTION_FIXBOOK_FILES = os.getenv("CORRECTION_FIXBOOK_FILES", "")
MEDICAL_TERMS_XLSX = os.getenv("MEDICAL_TERMS_XLSX", str(CORRECTION_DIR / "馬偕紀念醫院護理用語縮寫表.xlsx"))
ENABLE_MEDICAL_TERMS = os.getenv("ENABLE_MEDICAL_TERMS", "0") == "1"
USER_FIXBOOK_FILE = CORRECTION_DIR / "user_fixbook.json"

app = FastAPI(title="Meeting App Remote GPU Backend", version="0.1.0")

_tokenizer = None
_model = None
_whisper_model = None
_corrector = None


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    model: Optional[str] = None
    messages: List[ChatMessage]
    max_tokens: int = 1024
    temperature: float = 0.2
    stream: bool = False
    chat_template_kwargs: Optional[Dict[str, Any]] = None


class CorrectionRequest(BaseModel):
    wrong_word: str
    correct_word: str


def _load_llm():
    global _tokenizer, _model
    if _model is not None:
        return _tokenizer, _model
    _tokenizer = AutoTokenizer.from_pretrained(LLM_MODEL_ID, trust_remote_code=True)
    dtype = torch.float16 if torch.cuda.is_available() else torch.float32
    _model = AutoModelForCausalLM.from_pretrained(
        LLM_MODEL_ID,
        torch_dtype=dtype,
        device_map="auto" if torch.cuda.is_available() else None,
        trust_remote_code=True,
    )
    if not torch.cuda.is_available():
        _model.to("cpu")
    _model.eval()
    return _tokenizer, _model


def _load_whisper():
    global _whisper_model
    if _whisper_model is not None:
        return _whisper_model
    from faster_whisper import WhisperModel

    _whisper_model = WhisperModel(
        WHISPER_MODEL_SIZE,
        device=WHISPER_DEVICE,
        compute_type=WHISPER_COMPUTE_TYPE,
    )
    return _whisper_model


class TranscriptCorrector:
    def __init__(self, correction_dir: Path):
        self.replacements = []
        self.medical_terms = []
        self._load_fixbooks(correction_dir)
        if ENABLE_MEDICAL_TERMS:
            self._load_medical_terms(Path(MEDICAL_TERMS_XLSX))

    def _load_fixbooks(self, correction_dir: Path):
        pairs = {}
        names = [item.strip() for item in CORRECTION_FIXBOOK_FILES.split(",") if item.strip()]
        names.append(USER_FIXBOOK_FILE.name)
        for name in names:
            path = correction_dir / name
            if not path.exists():
                continue
            data = json.loads(path.read_text(encoding="utf-8"))
            for group_name in ("simple_replacements", "hybrid_corrections"):
                for item in data.get(group_name, []):
                    correct = str(item.get("correct_word") or "").strip()
                    if not correct:
                        continue
                    wrongs = list(item.get("wrong_words") or [])
                    wrongs.extend(item.get("phonetic_sounds") or [])
                    for wrong in wrongs:
                        wrong = str(wrong or "").strip()
                        if wrong and wrong != correct:
                            pairs[wrong] = correct
        self.replacements = sorted(pairs.items(), key=lambda pair: len(pair[0]), reverse=True)

    def add_replacement(self, wrong_word: str, correct_word: str):
        wrong_word = wrong_word.strip()
        correct_word = correct_word.strip()
        if not wrong_word or not correct_word:
            raise ValueError("wrong_word and correct_word are required")
        CORRECTION_DIR.mkdir(parents=True, exist_ok=True)
        data = {"simple_replacements": [], "hybrid_corrections": []}
        if USER_FIXBOOK_FILE.exists():
            data = json.loads(USER_FIXBOOK_FILE.read_text(encoding="utf-8"))
            data.setdefault("simple_replacements", [])
            data.setdefault("hybrid_corrections", [])
        rows = data["simple_replacements"]
        for row in rows:
            if row.get("correct_word") == correct_word:
                wrongs = row.setdefault("wrong_words", [])
                if wrong_word not in wrongs:
                    wrongs.append(wrong_word)
                break
        else:
            rows.append({"correct_word": correct_word, "wrong_words": [wrong_word]})
        USER_FIXBOOK_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        self.replacements = sorted(
            [(w, c) for (w, c) in self.replacements if w != wrong_word] + [(wrong_word, correct_word)],
            key=lambda pair: len(pair[0]),
            reverse=True,
        )

    def _load_medical_terms(self, path: Path):
        if not path.exists():
            return
        try:
            from openpyxl import load_workbook
        except Exception:
            return
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        terms = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            raw_eng = str(row[0] or "").strip()
            chi = str(row[2] or "").strip() if len(row) > 2 else ""
            if not raw_eng or not chi or "英文縮寫" in raw_eng:
                continue
            synonyms = [item.strip() for item in re.split(r"[()/\n]", raw_eng) if len(item.strip()) > 1]
            for synonym in synonyms:
                norm = re.sub(r"[^A-Za-z0-9]", "", synonym)
                if len(norm) < 2:
                    continue
                terms.append((synonym, norm, chi))
        self.medical_terms = sorted(terms, key=lambda item: len(item[1]), reverse=True)

    def correct(self, text: str) -> tuple[str, bool]:
        if not text:
            return text, False
        corrected = text
        for wrong, right in self.replacements:
            corrected = corrected.replace(wrong, right)
        corrected = self._annotate_medical_terms(corrected)
        return corrected, corrected != text

    def _annotate_medical_terms(self, text: str) -> str:
        corrected = text
        for synonym, norm, chi in self.medical_terms:
            if f"({chi})" in corrected:
                continue
            chars = [re.escape(ch) for ch in norm]
            flexible = r"[\s.\-_/]*".join(chars)
            if len(norm) <= 2:
                pattern = re.compile(rf"(?<![A-Za-z0-9])({flexible})(?![A-Za-z0-9])", re.IGNORECASE)
            else:
                pattern = re.compile(rf"(?<![A-Za-z0-9])({flexible})(?![A-Za-z0-9])", re.IGNORECASE)

            def repl(match):
                found = match.group(1)
                return f"{found} ({chi})"

            corrected = pattern.sub(repl, corrected)
        return corrected


def _load_corrector():
    global _corrector
    if _corrector is None:
        _corrector = TranscriptCorrector(CORRECTION_DIR)
    return _corrector


@app.get("/health")
def health():
    return {
        "ok": True,
        "llm_model": LLM_MODEL_ID,
        "whisper_model": WHISPER_MODEL_SIZE,
        "cuda": torch.cuda.is_available(),
        "gpu": torch.cuda.get_device_name(0) if torch.cuda.is_available() else None,
        "correction_dir": str(CORRECTION_DIR),
        "correction_loaded": _corrector is not None,
    }


@app.post("/debug/correct")
def debug_correct(payload: Dict[str, str]):
    text = payload.get("text", "")
    corrected, changed = _load_corrector().correct(text)
    return {"text": text, "corrected_text": corrected, "changed": changed}


@app.get("/corrections")
def list_corrections():
    if not USER_FIXBOOK_FILE.exists():
        return {"simple_replacements": []}
    data = json.loads(USER_FIXBOOK_FILE.read_text(encoding="utf-8"))
    return {"simple_replacements": data.get("simple_replacements", [])}


@app.post("/corrections/add")
def add_correction(req: CorrectionRequest):
    try:
        corrector = _load_corrector()
        corrector.add_replacement(req.wrong_word, req.correct_word)
        return {"ok": True, "wrong_word": req.wrong_word.strip(), "correct_word": req.correct_word.strip()}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/v1/chat/completions")
def chat_completions(req: ChatRequest):
    if req.stream:
        raise HTTPException(status_code=400, detail="stream=True is not supported by this minimal backend yet")
    tokenizer, model = _load_llm()
    messages = [{"role": m.role, "content": m.content} for m in req.messages]
    template_kwargs = req.chat_template_kwargs or {}
    if "enable_thinking" not in template_kwargs:
        template_kwargs["enable_thinking"] = False
    text = tokenizer.apply_chat_template(
        messages,
        tokenize=False,
        add_generation_prompt=True,
        **template_kwargs,
    )
    inputs = tokenizer([text], return_tensors="pt")
    inputs = {k: v.to(model.device) for k, v in inputs.items()}
    with torch.inference_mode():
        output_ids = model.generate(
            **inputs,
            max_new_tokens=req.max_tokens,
            temperature=req.temperature,
            do_sample=req.temperature > 0,
            pad_token_id=tokenizer.eos_token_id,
        )
    generated = output_ids[0][inputs["input_ids"].shape[-1] :]
    content = tokenizer.decode(generated, skip_special_tokens=True).strip()
    return {
        "id": "meeting-app-qwen",
        "object": "chat.completion",
        "model": req.model or LLM_MODEL_ID,
        "choices": [{"index": 0, "message": {"role": "assistant", "content": content}, "finish_reason": "stop"}],
    }


@app.post("/asr")
async def asr(
    file: UploadFile = File(...),
    language: str = Form("zh"),
    diarization: str = Form("0"),
):
    try:
        model = _load_whisper()
        suffix = os.path.splitext(file.filename or "")[1] or ".wav"
        data = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as handle:
            handle.write(data)
            path = handle.name
        segments, _info = model.transcribe(
            path,
            language=language or "zh",
            vad_filter=True,
            beam_size=5,
        )
        out_segments = []
        texts = []
        for seg in segments:
            text = (seg.text or "").strip()
            corrected_text, changed = _load_corrector().correct(text)
            texts.append(text)
            out_segments.append(
                {
                    "start": float(seg.start),
                    "end": float(seg.end),
                    "text": corrected_text,
                    "original_text": text if changed else "",
                    "speaker": "",
                }
            )
        raw_text = "".join(texts)
        corrected_full_text, text_changed = _load_corrector().correct(raw_text)
        if out_segments:
            corrected_full_text = "".join(item["text"] for item in out_segments)
        return {
            "text": corrected_full_text,
            "original_text": raw_text if text_changed else "",
            "segments": out_segments,
            "diarization": diarization == "1",
            "correction_applied": text_changed or any(item.get("original_text") for item in out_segments),
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"ASR failed: {exc}") from exc
    finally:
        if "path" in locals():
            try:
                os.remove(path)
            except OSError:
                pass
